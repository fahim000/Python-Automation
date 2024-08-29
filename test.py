import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Load the workbook and select the sheet based on the current day
workbook = openpyxl.load_workbook('4BeatsQ1.xlsx')
today = datetime.now().strftime("%A")
sheet = workbook[today]

# Initialize the Chrome WebDriver
driver = webdriver.Chrome()
driver.get('http://www.google.com')

# Function to perform the search and extract suggestions
def get_suggestions(keyword):
    print(f"Searching for: {keyword}")  # Debugging statement

    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(keyword)
    search_box.send_keys(Keys.RETURN)
    
    # Wait for search results to load
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "h3"))
        )
    except Exception as e:
        print(f"Error loading search results: {e}")
        return None, None

    # Capture the titles of the search results
    titles = driver.find_elements(By.TAG_NAME, "h3")
    suggestion_texts = [title.text for title in titles if title.text]

    print(f"Suggestions found: {suggestion_texts}")  # Debugging statement

    if suggestion_texts:
        longest = max(suggestion_texts, key=len)
        shortest = min(suggestion_texts, key=len)
        print(f"Longest: {longest}, Shortest: {shortest}")  # Debugging statement
        return longest, shortest
    else:
        print("No suggestions found.")  # Debugging statement
        return None, None

# Iterate through the keywords and update the Excel file
for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=3, max_col=3):
    keyword = row[0].value  # Fetching the keyword from column C

    print(f"Processing row: {row[0].row}, Keyword: {keyword}")  # Debugging statement

    if keyword:
        longest, shortest = get_suggestions(keyword)
        if longest and shortest:
            sheet.cell(row=row[0].row, column=4).value = longest  # Update column D with the longest suggestion
            sheet.cell(row=row[0].row, column=5).value = shortest  # Update column E with the shortest suggestion
        else:
            print(f"No suggestions were added for keyword: {keyword}")
    else:
        print(f"Empty cell found at row: {row[0].row}")

# Save the updated workbook and close the WebDriver
workbook.save('4BeatsQ1.xlsx')
driver.quit()
